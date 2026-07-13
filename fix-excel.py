import json
from openpyxl import load_workbook

print("Loading workbook...")
wb = load_workbook('NSPIRE-AI-.xlsx', read_only=True, data_only=True)
ws = wb.active

print("Reading headers...")
headers = []
for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
    headers = [str(h).strip() if h else '' for h in row]
    break

print(f"Headers: {headers}")

try:
    def_col_idx = headers.index('Deficiency Selected')
    proto_col_idx = headers.index('Inspection Protocol (International)')
except ValueError:
    print("Could not find required columns!")
    exit(1)

protocol_map = {}
print("Reading rows...")
for row in ws.iter_rows(min_row=2, values_only=True):
    def_sel = row[def_col_idx]
    proto = row[proto_col_idx]
    if def_sel and proto and str(proto).strip() and str(proto).strip() != 'None':
        protocol_map[str(def_sel).strip()] = str(proto).strip()

print(f"Found {len(protocol_map)} unique protocols mapping from deficiency names.")

json_path = 'lib/inspectionDeficiencies.json'
with open(json_path, 'r', encoding='utf-8') as f:
    data = json.load(f)

updated_count = 0
for section in ['outside', 'inside', 'unit']:
    for item in data.get(section, []):
        def_sel = item.get('deficiencySelected', '').strip()
        if def_sel in protocol_map:
            item['inspectionProtocol'] = protocol_map[def_sel]
            updated_count += 1
        elif not item.get('inspectionProtocol'):
            # Try partial match or lowercase
            for k, v in protocol_map.items():
                if def_sel.lower() == k.lower():
                    item['inspectionProtocol'] = v
                    updated_count += 1
                    break

with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(data, f, indent=2)

print(f"Successfully updated {updated_count} inspection protocols in JSON.")
